import re
from dataclasses import asdict
from typing import Optional, Dict, Iterator, Any

from openpyxl import Workbook
from openpyxl.cell import MergedCell, Cell
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet
from requests_html import HTML, Element

from jinja2xlsx.models import Style


def render(html_str: str, default_style: Optional[Style] = None) -> Workbook:
    default_style = default_style or Style()

    html = HTML(html=html_str)

    table = html.find("table", first=True)

    wb = Workbook()
    ws: Worksheet = wb.active

    colgroup = table.find("colgroup", first=True)
    if colgroup:
        adjust_columns(ws, table)

    table_body = table.find("tbody", first=True)
    fill_sheet_with_table_data(ws, table_body, default_style)

    return wb


def adjust_columns(sheet: Worksheet, colgroup: Element) -> None:
    columns = colgroup.find("col")
    for index, column in enumerate(columns):
        col_width_in_pixels = int(column.attrs.get("width", 0))
        if not col_width_in_pixels:
            continue

        column_dimension: ColumnDimension = sheet.column_dimensions[get_column_letter(index + 1)]
        column_dimension.width = width_pixels_to_xlsx_units(col_width_in_pixels)


def width_pixels_to_xlsx_units(pixels: float) -> float:
    return pixels / 7.5


def height_pixels_to_xlsx_units(pixels: float) -> float:
    return pixels * 3 / 4


def fill_sheet_with_table_data(sheet: Worksheet, table: Element, default_style: Style) -> None:
    row_index = 0
    col_index = 0

    for row in table.find("tr"):
        try_adjust_row(sheet, row_index, row)

        for html_cell in row.find("td"):
            target_cell = sheet.cell(row_index + 1, col_index + 1)
            while True:
                if isinstance(target_cell, MergedCell):
                    col_index += 1
                    target_cell = sheet.cell(row_index + 1, col_index + 1)
                else:
                    break

            target_cell.value = parse_cell_value(html_cell.text)

            colspan = int(html_cell.attrs.get("colspan", 1))
            rowspan = int(html_cell.attrs.get("rowspan", 1))

            style = extract_style(html_cell.attrs.get("style"))
            style = default_style.union(style)

            if colspan > 1 or rowspan > 1:
                cell_range = str_cell_range(
                    start_row=row_index + 1,
                    start_column=col_index + 1,
                    end_row=row_index + rowspan,
                    end_column=col_index + colspan,
                )
                style_and_merge_cell_range(sheet, cell_range, style)
            else:
                style_single_cell(target_cell, style)

            col_index += colspan

        row_index += 1
        col_index = 0


def try_adjust_row(sheet: Worksheet, row_index: int, row: Element) -> None:
    style_dict = style_to_dict(row.attrs.get("style"))
    height_str = style_dict.get("line-height") or style_dict.get("height") or ""
    row_height = try_extract_pixels(height_str)
    if row_height:
        sheet.row_dimensions[row_index + 1].height = height_pixels_to_xlsx_units(row_height)


def try_extract_pixels(pixel_str: Optional[str]) -> Optional[float]:
    """
    >>> try_extract_pixels("100px")
    100.0
    >>> try_extract_pixels("") is None
    True
    """
    if not pixel_str:
        return None

    return float(re.findall("(\d+)px", pixel_str)[0])


def parse_cell_value(cell_text: str) -> Any:
    """
    >>> parse_cell_value("") is None
    True
    >>> parse_cell_value("ass")
    'ass'
    >>> parse_cell_value("1")
    1
    >>> parse_cell_value("1.2")
    1.2
    """
    try:
        return int(cell_text)
    except ValueError:
        try:
            return float(cell_text)
        except ValueError:
            if cell_text == "":
                return None

            return cell_text


def extract_style(style_attr: str) -> Style:
    """
    >>> style = extract_style("border: 1px solid black; text-align: center; font-weight: bold")
    >>> style.alignment.horizontal
    'center'
    >>> style.border.left.style
    'thin'
    >>> style.border.left.style == style.border.right.style == style.border.top.style == style.border.bottom.style
    True
    >>> style.font.bold
    True
    """
    if not style_attr:
        return Style()

    style_dict = style_to_dict(style_attr)

    border = _build_border(style_dict)
    alignment = _build_alignment(style_dict)
    font = _build_font(style_dict)

    return Style(border, alignment, font)


def style_to_dict(style_str: Optional[str]) -> Dict:
    """
    >>> style_to_dict("border: 1px solid black; text-align: center; font-weight: bold")
    {'border': '1px solid black', 'text-align': 'center', 'font-weight': 'bold'}
    >>> style_to_dict("")
    {}
    >>> style_to_dict(None)
    {}
    """
    if not style_str:
        return {}

    return {
        style.strip(): value.strip()
        for style, value in (style.split(":") for style in filter(None, style_str.split(";")))
    }


def _build_border(style_dict: Dict[str, str]) -> Border:
    """
    >>> border = _build_border({"border": "1px solid black"})
    >>> border.left.style
    'thin'
    >>> border.left.style == border.right.style == border.top.style == border.bottom.style
    True
    >>> border = _build_border({"border-right": "2px solid black"})
    >>> border.right.style
    'medium'
    """

    def _from_border_attr(border_attr: str) -> Optional[Border]:
        border_rule = style_dict.get(border_attr)
        if not border_rule:
            return None

        if border_rule == "1px solid black":
            side = Side(style="thin")
        elif re.match(r"\d+px solid black", border_rule):
            side = Side(style="medium")
        else:
            side = Side()

        if border_attr == "border":
            return Border(left=side, right=side, top=side, bottom=side)
        if border_attr == "border-left":
            return Border(left=side)
        if border_attr == "border-right":
            return Border(right=side)
        if border_attr == "border-top":
            return Border(top=side)
        if border_attr == "border-bottom":
            return Border(bottom=side)

        return None

    borders: Iterator[Border] = filter(
        None,
        (
            _from_border_attr("border"),
            _from_border_attr("border-left"),
            _from_border_attr("border-right"),
            _from_border_attr("border-top"),
            _from_border_attr("border-bottom"),
        ),
    )

    return next(borders, Border())


def _build_alignment(style_dict: Dict) -> Alignment:
    word_wrap = style_dict.get("word-wrap")

    wrap_text: Optional[bool]
    if word_wrap == "break-word":
        wrap_text = True
    elif word_wrap == "normal":
        wrap_text = False
    else:
        wrap_text = None

    alignment = Alignment(horizontal=style_dict.get("text-align"), wrap_text=wrap_text)
    return alignment


def _build_font(style_dict: Dict) -> Font:
    font = Font(bold=style_dict.get("font-weight") == "bold")
    return font


def style_single_cell(cell: Cell, style: Style) -> None:
    for style_key, value in asdict(style).items():
        setattr(cell, style_key, value)


def str_cell_range(start_column: int, start_row: int, end_column: int, end_row: int) -> str:
    from_column = get_column_letter(start_column)
    to_column = get_column_letter(end_column)
    return f"{from_column}{start_row}:{to_column}{end_row}"


def style_and_merge_cell_range(sheet: Worksheet, cell_range: str, style: Style) -> None:
    """
    Source:
    https://openpyxl.readthedocs.io/en/2.5/styles.html#styling-merged-cells
    """
    top = Border(top=style.border.top)
    left = Border(left=style.border.left)
    right = Border(right=style.border.right)
    bottom = Border(bottom=style.border.bottom)

    first_cell = sheet[cell_range.split(":")[0]]
    if style.alignment:
        sheet.merge_cells(cell_range)
        first_cell.alignment = style.alignment

    if style.font:
        first_cell.font = style.font

    rows = sheet[cell_range]
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
